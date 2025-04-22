import openpyxl
import os

#if there no sheet then create a new sheet
def create_attendance_file(excel_path):
    if not os.path.exists(excel_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance_Record"
        sheet.cell(row=1, column=1, value="Student_ID")
        for student_num in range(1, 101):
            sheet.cell(row=student_num+1, column=1, value=student_num)
        workbook.save(excel_path)

def mark_attendance(excel_path, present_students, lecture_date):
    create_attendance_file(excel_path)
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    #check for the today's date column is present or not
    today_column = None
    for col_idx in range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col_idx).value == lecture_date:
            today_column = col_idx
            break

    # Create new date column if needed
    if not today_column:
        today_column = sheet.max_column + 1
        sheet.cell(row=1, column=today_column, value=lecture_date)

    # Update attendance status for all students 
    for student_row in range(2, sheet.max_row + 1):
        student_id = sheet.cell(row=student_row, column=1).value
        if str(student_id) in present_students:
            sheet.cell(row=student_row, column=today_column, value='P')
        else:
            sheet.cell(row=student_row, column=today_column, value='A')

    workbook.save(excel_path)